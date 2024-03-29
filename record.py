import telebot
import pytz
import os
import openpyxl
import schedule
import time
from telebot import types
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from datetime import datetime, timedelta

load_dotenv()

TOKEN = os.getenv('token')

bot = telebot.TeleBot(TOKEN)

file_path = 'data/data.xlsx'

busy_dates = []

user_dict = {}


def load_busy_slots() -> list:
    """
    Функция для формирования списка с занятыми датами

    :return: Список с занятыми датами
    """
    global busy_dates
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        date_column = sheet['C']
        busy_dates = [str(cell.value) for cell in date_column if cell.value is not None]

        return busy_dates
    except Exception as e:
        print(f"Error loading busy_dates: {e}")


def load_user_dict() -> dict:
    """
    Формирование словаря для напоминания о записи!
    """
    global user_dict

    book = openpyxl.load_workbook(file_path)
    sheet = book.active

    date = sheet['C']
    user = sheet['D']

    user_date = [str(cell.value) for cell in date if cell.value is not None]
    user_id = [str(cell.value) for cell in user if cell.value is not None]

    user_dict = {date_value: id_value for date_value, id_value in zip(user_date, user_id)}

    return user_dict


def save_to_excel(data):
    """
    Функция для записи в файл данных о клиенте и дате записи
    """
    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook.active

    sheet.insert_rows(1)

    for col_num, value in enumerate(data, start=1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = value

    workbook.save(file_path)


@bot.message_handler(commands=['start'])
def handler_start(message):
    text_mess = "Привет! Я бот для записи, напишите пожалуйста ваше имя"
    bot.send_message(message.chat.id, text_mess)
    bot.register_next_step_handler(message, handle_name)


@bot.message_handler(commands=['file'])
def handler_get_file():
    """
    Функция для запроса файла с данными о клиентах
    """
    chat_id = os.getenv('user_file')
    file = open('data/data.xlsx', 'rb')
    bot.send_document(chat_id, file)


def handle_name(message):
    user_data = {'name': message.text}
    text_mess = "Укажите номер телефона"
    bot.send_message(message.chat.id, text_mess)
    bot.register_next_step_handler(message, handler_phone, user_data)


def handler_phone(message, user_data):
    phone_number = message.text
    if phone_number.isdigit() and len(phone_number) == 11:
        user_data['phone'] = phone_number
        load_busy_slots()
        show_available_slots(message.chat.id)
        bot.register_next_step_handler(message, handler_date_time, user_data)
    else:
        bot.send_message(message.chat.id, "Проверьте правильность телефонного номера")
        bot.register_next_step_handler(message, show_available_slots, user_data)


@bot.message_handler(func=lambda message: message.text in get_available_slots())
def handler_date_time(message, user_data):
    selected_date = message.text
    if selected_date in get_available_slots():
        bot.send_message(message.chat.id, f"Вы выбрали дату: {selected_date}")
        user_data['date_time'] = selected_date
        user_data['user_id'] = message.from_user.id
        save_to_excel([user_data['name'], user_data['phone'], user_data['date_time'], user_data['user_id']])
        send_message_boss(user_data)
        text_mess = (f"Вы успешно записаны\n"
                     f"Дата и время приёма: {user_data['date_time']}")

        # Убираем кнопки после выбора пользователя
        close_button = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, text_mess, reply_markup=close_button)

    else:
        bot.send_message(message.chat.id, 'Выберите дату из списка')
        bot.register_next_step_handler(message, handler_date_time, user_data)


def generate_all_slots() -> list:
    """
    Функция для формирования списка слотов

    :return: Список слотов
    """

    local_zone = pytz.timezone('Asia/Yakutsk')
    start_time = datetime.now(local_zone)

    time_slots = ['9:15', '12:00', '15:00']

    if start_time.hour > 15:
        start_time += timedelta(days=1)

    num_days = 31

    all_slots = []

    for day in range(num_days):
        currents_date = start_time + timedelta(days=day)
        for time_slot in time_slots:
            slot_datetime_str = f"{currents_date.strftime('%d-%m-%Y')} {time_slot}"
            all_slots.append(slot_datetime_str)

    if 9 <= start_time.hour < 12:
        del all_slots[0]

    if 12 <= start_time.hour < 15:
        del all_slots[0]
        del all_slots[0]

    return all_slots


def get_filter_all_slots() -> list:
    """
    Функция для формирования отфильтрованного списка(исключение Воскресенья)

    :return: Отфильтрованый список
    """
    no_filter = generate_all_slots()

    filter_slots = []
    for no_filters in no_filter:
        date_obj = datetime.strptime(no_filters, '%d-%m-%Y %H:%M')
        if date_obj.weekday() != 6:
            str_obj = date_obj.strftime('%d-%m-%Y %H:%M')
            filter_slots.append(str_obj)

    return filter_slots


def get_available_slots() -> list:
    """
     Функция для формирования свободных дат для записи

    :return: Список свободных дат
    """
    all_slots = get_filter_all_slots()
    available_slots = []

    for slot in all_slots:

        if slot not in busy_dates:
            available_slots.append(slot)

    return available_slots


@bot.message_handler(commands=['slots'])
def show_available_slots(chat_id):
    """
    Функция для предоставления пользователю список свободных дат
    """
    slots = get_available_slots()

    if not slots:
        bot.send_message(chat_id, f"Извините, все даты заняты")
        return

    # Создаём кнопки с свободными слотами
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    for slot in slots:
        button = types.KeyboardButton(text=slot)
        keyboard.add(button)

    bot.send_message(chat_id, "Выберите удобную дату:", reply_markup=keyboard)


def send_message_boss(user_data):
    """
    Функция для отправки данных о новой записи на приём
    """
    user_id = os.getenv('user_message')
    bot.send_message(user_id, f"Новая запись: {user_data['name']}\n{user_data['phone']}\n{user_data['date_time']}")


def schedule_reminder():
    user_data = load_user_dict()

    for date in user_data.keys():
        reminder_time = datetime.strptime(date, '%d-%m-%Y %H:%M') - timedelta(hours=11, minutes=30)
        schedule.every().day.at(reminder_time.strftime('%H:%M')).do(send_reminder, date)


def send_reminder(date):
    data = load_user_dict()

    user_id = data[date]

    bot.send_message(user_id, f"Напоминание: У вас запись через 24 часа на {date}")


def main():
    bot.polling(none_stop=True)
    schedule_reminder()
    while True:
        schedule.run_pending()
        time.sleep(1)


if __name__ == '__main__':
    main()
